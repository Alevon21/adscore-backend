"""
FastAPI application for Haraba Text Scoring prototype.
All data is stored in-memory with a 60-minute TTL per session.
"""

from dotenv import load_dotenv
load_dotenv(override=True)

import hashlib
import io
import ipaddress
import logging
import os
import threading
import uuid as uuid_mod
from datetime import datetime, timezone
from typing import Any, Dict
from urllib.parse import urlparse

import pandas as pd
from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from exporter import export_results
from mapper import ColumnMapper
from campaign_scorer import analyze_campaigns
from models import (
    ABTestRequest,
    CampaignAnalysisRequest,
    EventConfig,
    ExtractedWord,
    ExtractWordsRequest,
    ExtractWordsResult,
    MappingRequest,
    ScoreRequest,
    ScoringParams,
    ScoringResult,
    TextPartRequest,
)
from scorer import TextScorer
from tester import ABTester
from text_analyzer import TextPartAnalyzer
from adscore import adscore_router
from users import router as auth_router, tenant_router, log_audit
from auth import get_current_user, CurrentUser
from sessions import router as sessions_router
from usability_test import router as usability_test_router
from mmp import router as mmp_router
from demo import router as demo_router
from creative_history import creative_history_router
from creative_analytics import analytics_router
from share import router as share_router
from comments import router as comments_router
from activity_feed import router as activity_router
from competitor_analysis import router as competitor_router
from predictive_ctr import router as predictive_ctr_router
from creative_generator import router as creative_generator_router
from cross_platform import router as cross_platform_router
from trend_detection import router as trend_router
from database import get_db
from db_models import (
    ScoringSession, ScoringResult as ScoringResultDB,
    StoredFile, FileStatus, SessionStatus,
)
import storage as file_storage

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="Haraba Text Scoring",
    description="Local prototype for marketing text scoring",
    version="2.0.0",
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.include_router(adscore_router)
app.include_router(auth_router)
app.include_router(tenant_router)
app.include_router(sessions_router)
app.include_router(usability_test_router)
app.include_router(mmp_router)
app.include_router(demo_router)
app.include_router(creative_history_router)
app.include_router(analytics_router)
app.include_router(share_router)
app.include_router(comments_router)
app.include_router(activity_router)
app.include_router(competitor_router)
app.include_router(predictive_ctr_router)
app.include_router(creative_generator_router)
app.include_router(cross_platform_router)
app.include_router(trend_router)

_cors_origins = [
    "http://localhost:3000",
    "http://localhost:3001",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:3001",
    "http://frontend:3000",
    "https://adscore-orpin.vercel.app",
]
_frontend_url = os.getenv("FRONTEND_URL", "")
if _frontend_url:
    _cors_origins.append(_frontend_url)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Requested-With"],
)


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response

@app.on_event("startup")
async def on_startup():
    from database import init_db
    try:
        await init_db()
        logger.info("Database initialized")
    except Exception as e:
        logger.warning("Database init skipped (not available): %s", e)


# In-memory session store (protected by _session_lock)
SESSION_STORE: Dict[str, Dict[str, Any]] = {}
SESSION_TIMERS: Dict[str, threading.Timer] = {}
_session_lock = threading.Lock()
SESSION_TTL = int(os.getenv("SESSION_TTL_MINUTES", "60")) * 60  # seconds
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE_MB", "50")) * 1024 * 1024  # bytes

mapper = ColumnMapper()

# ── SSRF protection ──────────────────────────────────────────────
_SSRF_BLOCKED_HOSTS = {
    'localhost', '127.0.0.1', '0.0.0.0', '::1',
    '169.254.169.254', 'metadata.google.internal',
    'metadata.internal', '100.100.100.200',
}


def _is_safe_url(url: str) -> bool:
    """Validate URL is safe for server-side fetch (blocks SSRF)."""
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    if parsed.scheme not in ('http', 'https'):
        return False
    hostname = parsed.hostname
    if not hostname:
        return False
    if hostname in _SSRF_BLOCKED_HOSTS:
        return False
    try:
        ip = ipaddress.ip_address(hostname)
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
            return False
    except ValueError:
        pass  # hostname is a domain name, not IP — OK
    return True


def _cleanup_session(session_id: str) -> None:
    """Remove a session after TTL expires (called from Timer thread)."""
    with _session_lock:
        SESSION_STORE.pop(session_id, None)
        SESSION_TIMERS.pop(session_id, None)
    logger.info("Session %s expired and cleaned up", session_id)


def _schedule_ttl_locked(session_id: str) -> None:
    """Reset the TTL timer. Caller MUST hold _session_lock."""
    old_timer = SESSION_TIMERS.pop(session_id, None)
    if old_timer:
        old_timer.cancel()
    timer = threading.Timer(SESSION_TTL, _cleanup_session, args=[session_id])
    timer.daemon = True
    timer.start()
    SESSION_TIMERS[session_id] = timer


def _get_session(session_id: str, tenant_id: uuid_mod.UUID = None) -> Dict[str, Any]:
    """Retrieve session data or raise 404. Verifies tenant ownership if tenant_id given."""
    with _session_lock:
        if session_id not in SESSION_STORE:
            raise HTTPException(status_code=404, detail="Session not found")
        data = SESSION_STORE[session_id]
        if tenant_id is not None and data.get("tenant_id") != tenant_id:
            raise HTTPException(status_code=404, detail="Session not found")
        _schedule_ttl_locked(session_id)
        return data


# ---------- ENDPOINTS ----------


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/upload")
@limiter.limit("20/minute")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Upload CSV or XLSX file.
    Returns session_id, detected columns, auto-mapping, event info, and scoring mode.
    """
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format: .{ext}. Use CSV or XLSX.",
        )

    # Stream-read with size limit to avoid loading huge files into memory
    chunks = []
    total = 0
    while True:
        chunk = await file.read(1024 * 1024)  # 1 MB chunks
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Max size: {MAX_FILE_SIZE // (1024*1024)}MB",
            )
        chunks.append(chunk)
    content = b"".join(chunks)

    try:
        buf = io.BytesIO(content)
        if ext == "csv":
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf, engine="openpyxl")
    except Exception as e:
        logger.error("Failed to parse file: %s", e)
        raise HTTPException(status_code=400, detail="Failed to parse file. Please check the format and try again.")

    if df.empty:
        raise HTTPException(status_code=400, detail="File is empty")

    columns_detected = list(df.columns.astype(str))
    auto_mapped = mapper.auto_map(columns_detected)
    unmapped = mapper.get_unmapped_columns(columns_detected, auto_mapped)

    # Auto-detect events from mapping
    auto_events = mapper.detect_events(auto_mapped)

    # Detect mode based on auto-mapped columns
    temp_df = mapper.apply_mapping(df.copy(), auto_mapped)
    event_configs = [EventConfig(**e) for e in auto_events]
    scorer_tmp = TextScorer(ScoringParams(events=event_configs))
    mode = scorer_tmp.detect_mode(temp_df)

    session_id = str(uuid_mod.uuid4())
    with _session_lock:
        SESSION_STORE[session_id] = {
            "tenant_id": current_user.tenant.id,
            "df_original": df,
            "df_mapped": None,
            "columns_detected": columns_detected,
            "auto_mapped": auto_mapped,
            "mapping": auto_mapped.copy(),
            "events": auto_events,
            "mode": mode,
            "scoring_result": None,
            "params": None,
            "text_part_result": None,
        }
        _schedule_ttl_locked(session_id)

    logger.info(
        "Upload: session=%s, rows=%d, cols=%d, mode=%s, events=%d",
        session_id, len(df), len(columns_detected), mode, len(auto_events),
    )

    # Persist to DB + Supabase Storage
    from database import async_session
    async with async_session() as db:
        # Audit log
        await log_audit(
            db, current_user.tenant.id, current_user.user.id, "upload", request,
            resource_type="file", resource_id=session_id,
            details={"filename": filename, "rows": len(df), "mode": mode},
        )

        # Dual-write: persist session + file to DB
        try:
            sid = uuid_mod.UUID(session_id)
            tid = current_user.tenant.id
            uid = current_user.user.id
            mime = "text/csv" if ext == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            # Try Supabase Storage upload (non-blocking for DB persistence)
            storage_key = ""
            file_id = None
            try:
                await file_storage.check_storage_quota(db, tid, len(content))
                storage_key = await file_storage.upload_file(tid, sid, filename, content, mime)

                stored_file = StoredFile(
                    tenant_id=tid, original_name=filename, storage_key=storage_key,
                    size_bytes=len(content), mime_type=mime,
                    checksum_sha256=hashlib.sha256(content).hexdigest(),
                    status=FileStatus.ready, uploaded_by=uid,
                )
                db.add(stored_file)
                await db.flush()
                file_id = stored_file.id
            except HTTPException:
                raise  # re-raise quota exceeded
            except Exception as e:
                logger.warning("Storage upload failed for session %s (non-fatal): %s", session_id, e)

            # Create ScoringSession row (even if storage upload failed)
            db_session = ScoringSession(
                id=sid, tenant_id=tid, file_id=file_id,
                status=SessionStatus.uploaded, mode=mode, n_rows=len(df),
                file_name=filename, columns_detected=columns_detected,
                auto_mapped=auto_mapped, events_detected=auto_events,
                created_by=uid,
            )
            db.add(db_session)
            await db.commit()
            logger.info("Session %s persisted to DB (file_id=%s)", session_id, file_id)
        except HTTPException:
            raise
        except Exception as e:
            logger.error("Failed to persist session %s to DB: %s", session_id, e)
            # Don't fail the upload — in-memory session still works

    return {
        "session_id": session_id,
        "columns_detected": columns_detected,
        "auto_mapped": auto_mapped,
        "unmapped": unmapped,
        "n_rows": len(df),
        "mode_detected": mode,
        "events_detected": auto_events,
    }


@app.post("/map")
async def apply_mapping(req: MappingRequest, current_user: CurrentUser = Depends(get_current_user)):
    """
    Apply user-defined column mapping to the uploaded data.
    Accepts optional event configs.
    """
    session = _get_session(req.session_id, current_user.tenant.id)

    is_valid, missing = mapper.validate_mapping(req.mapping)
    if not is_valid:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required fields: {', '.join(missing)}",
        )

    df_original = session["df_original"]
    df_mapped = mapper.apply_mapping(df_original.copy(), req.mapping)

    # Use provided events or auto-detect
    if req.events:
        events = [e.model_dump() for e in req.events]
    else:
        events = mapper.detect_events(req.mapping)

    # Rename event columns from original names to slot names (e.g. conversions_1 → event_1)
    for ev in events:
        col = ev.get("column")
        slot = ev.get("slot")
        if col and slot and col in df_mapped.columns and col != slot:
            df_mapped = df_mapped.rename(columns={col: slot})

    event_configs = [EventConfig(**e) for e in events]
    scorer_tmp = TextScorer(ScoringParams(events=event_configs))
    mode = scorer_tmp.detect_mode(df_mapped)

    session["df_mapped"] = df_mapped
    session["mapping"] = req.mapping
    session["events"] = events
    session["mode"] = mode

    logger.info("Mapping applied for session %s, mode=%s, events=%d", req.session_id, mode, len(events))

    # Persist mapping to DB
    try:
        from database import async_session
        async with async_session() as db:
            sid = uuid_mod.UUID(req.session_id)
            await db.execute(
                update(ScoringSession)
                .where(ScoringSession.id == sid)
                .values(mapping=req.mapping, events_detected=events, status=SessionStatus.mapped, mode=mode)
            )
            await db.commit()
    except Exception as e:
        logger.error("Failed to persist mapping for session %s: %s", req.session_id, e)

    return {"status": "ok", "mode": mode, "events": events}


@app.post("/score")
@limiter.limit("30/minute")
async def run_scoring(req: ScoreRequest, request: Request, current_user: CurrentUser = Depends(get_current_user)):
    """
    Run the full scoring pipeline on the mapped data.
    """
    session = _get_session(req.session_id, current_user.tenant.id)

    df_mapped = session.get("df_mapped")
    if df_mapped is None:
        current_mapping = session.get("mapping", session.get("auto_mapped", {}))
        if not current_mapping:
            raise HTTPException(
                status_code=400,
                detail="No mapping applied. Call /map first.",
            )
        df_mapped = mapper.apply_mapping(
            session["df_original"].copy(), current_mapping
        )
        # Rename event columns in fallback path too
        session_events_fb = session.get("events", [])
        for ev in session_events_fb:
            col = ev.get("column")
            slot = ev.get("slot")
            if col and slot and col in df_mapped.columns and col != slot:
                df_mapped = df_mapped.rename(columns={col: slot})

    # Inject event configs from session if not in params
    params = req.params
    if not params.events:
        session_events = session.get("events", [])
        if session_events:
            params = params.model_copy(update={
                "events": [EventConfig(**e) for e in session_events]
            })

    scorer = TextScorer(params)
    result: ScoringResult = scorer.score(df_mapped)

    session["scoring_result"] = result
    session["params"] = params

    logger.info(
        "Scoring done: session=%s, scored=%d, mode=%s",
        req.session_id, result.stats.get("n_scored", 0), result.stats.get("mode"),
    )

    result_dump = result.model_dump()

    # Persist to DB
    from database import async_session
    async with async_session() as db:
        await log_audit(
            db, current_user.tenant.id, current_user.user.id, "score", request,
            resource_type="session", resource_id=req.session_id,
            details={"n_scored": result.stats.get("n_scored", 0), "mode": result.stats.get("mode")},
        )

        try:
            sid = uuid_mod.UUID(req.session_id)
            tid = current_user.tenant.id
            params_dump = params.model_dump() if params else None

            # Update session status
            await db.execute(
                update(ScoringSession)
                .where(ScoringSession.id == sid)
                .values(
                    status=SessionStatus.completed,
                    params=params_dump,
                    completed_at=datetime.now(timezone.utc),
                )
            )

            # Save scoring results
            db_result = ScoringResultDB(
                session_id=sid, tenant_id=tid,
                results=result_dump.get("results"),
                stats=result_dump.get("stats"),
            )
            db.add(db_result)
            await db.commit()
            logger.info("Scoring results persisted for session %s", req.session_id)
        except Exception as e:
            logger.error("Failed to persist scoring results for %s: %s", req.session_id, e)

    return result_dump


@app.post("/process-banners")
@limiter.limit("10/minute")
async def process_banners(
    request: Request,
    req: dict,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Process banner_url column from CSV — download images, create Banner records."""
    import asyncio
    from adscore import create_banner_from_url

    session_id = req.get("session_id", "")
    with _session_lock:
        session = SESSION_STORE.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.get("tenant_id") != current_user.tenant.id:
        raise HTTPException(status_code=404, detail="Session not found")

    # Idempotency guard
    if session.get("banner_ids"):
        return {"processed": len(session["banner_ids"]), "failed": 0, "banner_ids": session["banner_ids"]}

    df = session.get("df_mapped", session.get("df_original"))
    if df is None or "banner_url" not in df.columns:
        return {"processed": 0, "failed": 0, "banner_ids": {}}

    tid = current_user.tenant.id
    uid = current_user.user.id

    # Collect rows with banner_url
    rows_with_banners = []
    for idx, row in df.iterrows():
        url = row.get("banner_url")
        if not url or not isinstance(url, str) or not url.strip():
            continue
        url = url.strip()
        if not url.startswith("http"):
            continue
        if not _is_safe_url(url):
            continue

        # Build metrics dict from CSV columns — store ALL raw data
        metrics_dict = {}

        # Raw integer counts
        for field in ("impressions", "clicks", "installs"):
            val = _safe_int(row.get(field))
            if val is not None:
                metrics_dict[field] = val

        # Raw float values
        for field in ("spend", "revenue"):
            val = _safe_float(row.get(field))
            if val is not None:
                metrics_dict[field] = val

        # Funnel events (up to 4)
        for i in range(1, 5):
            val = _safe_int(row.get(f"event_{i}"))
            if val is not None:
                metrics_dict[f"event_{i}"] = val

        # Computed rates
        impressions = metrics_dict.get("impressions")
        clicks = metrics_dict.get("clicks")
        if impressions and clicks:
            metrics_dict["ctr"] = clicks / impressions
        installs = metrics_dict.get("installs")
        if installs is not None and clicks:
            metrics_dict["cr_install"] = installs / clicks

        # Metadata
        for field in ("platform", "campaign", "date_from", "date_to"):
            val = row.get(field)
            if val is not None and str(val).strip():
                metrics_dict[field] = str(val).strip()

        # Event labels from session events
        events = session.get("events", [])
        if events:
            event_labels = {e["slot"]: e["label"] for e in events if e.get("label")}
            if event_labels:
                metrics_dict["event_labels"] = event_labels

        text_id = str(row.get("text_id", idx))
        rows_with_banners.append((text_id, url, metrics_dict))

    if not rows_with_banners:
        return {"processed": 0, "failed": 0, "banner_ids": {}}

    # Process sequentially to avoid concurrent AsyncSession usage
    banner_ids = {}
    failed = 0

    for text_id, banner_url, metrics in rows_with_banners:
        bid = await create_banner_from_url(db, tid, uid, banner_url, metrics)
        if bid:
            banner_ids[text_id] = str(bid)
        else:
            failed += 1

    await db.commit()

    session["banner_ids"] = banner_ids
    logger.info("Processed %d banners for session %s (%d failed)", len(banner_ids), session_id, failed)

    return {"processed": len(banner_ids), "failed": failed, "banner_ids": banner_ids}


def _safe_int(val):
    """Safely convert a value to int, returning None on failure."""
    if val is None:
        return None
    try:
        v = float(val)
        if pd.isna(v):
            return None
        return int(v)
    except (ValueError, TypeError):
        return None


def _safe_float(val):
    """Safely convert a value to float, returning None on failure."""
    if val is None:
        return None
    try:
        v = float(val)
        if pd.isna(v):
            return None
        return v
    except (ValueError, TypeError):
        return None


@app.post("/abtest")
def run_abtest(req: ABTestRequest, current_user: CurrentUser = Depends(get_current_user)):
    """
    Run A/B test comparing two texts from the current session.
    """
    session = _get_session(req.session_id, current_user.tenant.id)
    scoring_result = session.get("scoring_result")
    if scoring_result is None:
        raise HTTPException(
            status_code=400,
            detail="No scoring results. Run /score first.",
        )

    result_a = None
    result_b = None
    for r in scoring_result.results:
        if r.text_id == req.text_id_a:
            result_a = r
        if r.text_id == req.text_id_b:
            result_b = r

    if result_a is None:
        raise HTTPException(status_code=404, detail=f"Text {req.text_id_a} not found")
    if result_b is None:
        raise HTTPException(status_code=404, detail=f"Text {req.text_id_b} not found")

    params = session.get("params", ScoringParams())
    tester = ABTester()
    comparison = tester.compare(
        result_a.model_dump(),
        result_b.model_dump(),
        req.metric,
        fdr_level=params.fdr_level if params else 0.01,
    )

    logger.info(
        "A/B test: %s vs %s on %s, p=%.4f",
        req.text_id_a, req.text_id_b, req.metric,
        comparison.get("p_value", 1),
    )

    return comparison


@app.post("/text-parts")
async def run_text_parts(req: TextPartRequest, current_user: CurrentUser = Depends(get_current_user)):
    """
    Analyze text parts (elements) and find best combinations.
    """
    session = _get_session(req.session_id, current_user.tenant.id)
    scoring_result = session.get("scoring_result")
    if scoring_result is None:
        raise HTTPException(
            status_code=400,
            detail="No scoring results. Run /score first.",
        )

    # Get all headlines for auto-detection (including texts excluded by scoring filters)
    df_mapped = session.get("df_mapped")
    all_headlines = None
    if df_mapped is not None and "headline" in df_mapped.columns:
        all_headlines = df_mapped["headline"].fillna("").astype(str).tolist()

    params = session.get("params", ScoringParams())
    fdr_level = params.fdr_level if params else 0.01

    analyzer = TextPartAnalyzer()
    result = analyzer.analyze(
        scoring_result=scoring_result,
        custom_parts=req.custom_parts,
        primary_metric=req.primary_metric,
        max_combination_size=req.max_combination_size,
        all_headlines=all_headlines,
        fdr_level=fdr_level,
    )

    session["text_part_result"] = result

    logger.info(
        "Text parts analysis: session=%s, parts=%d",
        req.session_id, len(result.parts_detected),
    )

    text_part_dump = result.model_dump()

    # Persist to DB
    try:
        from database import async_session
        async with async_session() as db:
            sid = uuid_mod.UUID(req.session_id)
            await db.execute(
                update(ScoringResultDB)
                .where(ScoringResultDB.session_id == sid)
                .values(text_part_result=text_part_dump)
            )
            await db.commit()
    except Exception as e:
        logger.error("Failed to persist text-parts for %s: %s", req.session_id, e)

    return text_part_dump


@app.post("/extract-words")
def extract_words(req: ExtractWordsRequest, current_user: CurrentUser = Depends(get_current_user)):
    """
    Extract all unique words from headlines for interactive selection.
    """
    session = _get_session(req.session_id, current_user.tenant.id)
    scoring_result = session.get("scoring_result")
    if scoring_result is None:
        raise HTTPException(
            status_code=400,
            detail="No scoring results. Run /score first.",
        )

    scored_headlines = [t.headline for t in scoring_result.results]

    analyzer = TextPartAnalyzer()
    words = analyzer.extract_all_words(
        scored_headlines,
        min_length=req.min_length,
        include_bigrams=req.include_bigrams,
    )

    logger.info(
        "Extract words: session=%s, words=%d",
        req.session_id, len(words),
    )

    return ExtractWordsResult(
        words=words,
        n_texts=len(scored_headlines),
        headlines=scored_headlines,
    ).model_dump()


@app.post("/campaign-analysis")
async def run_campaign_analysis(req: CampaignAnalysisRequest, current_user: CurrentUser = Depends(get_current_user)):
    """Analyze campaign-level performance from scored results."""
    session = _get_session(req.session_id, current_user.tenant.id)
    scoring_result = session.get("scoring_result")
    if scoring_result is None:
        raise HTTPException(
            status_code=400,
            detail="No scoring results. Run /score first.",
        )

    campaigns_in_data = {r.campaign for r in scoring_result.results if r.campaign}
    if len(campaigns_in_data) < 2:
        raise HTTPException(
            status_code=400,
            detail=f"Need at least 2 campaigns for analysis. Found: {len(campaigns_in_data)}.",
        )

    params = session.get("params", ScoringParams())
    df_mapped = session.get("df_mapped")

    result = analyze_campaigns(
        results=scoring_result.results,
        df_mapped=df_mapped,
        params=params,
        stats=scoring_result.stats,
    )

    session["campaign_analysis"] = result

    logger.info(
        "Campaign analysis: session=%s, campaigns=%d",
        req.session_id, result.n_campaigns,
    )

    campaign_dump = result.model_dump()

    # Persist to DB
    try:
        from database import async_session
        async with async_session() as db:
            sid = uuid_mod.UUID(req.session_id)
            await db.execute(
                update(ScoringResultDB)
                .where(ScoringResultDB.session_id == sid)
                .values(campaign_analysis=campaign_dump)
            )
            await db.commit()
    except Exception as e:
        logger.error("Failed to persist campaign-analysis for %s: %s", req.session_id, e)

    return campaign_dump


@app.get("/export/{session_id}")
async def export_xlsx(session_id: str, request: Request, current_user: CurrentUser = Depends(get_current_user)):
    """
    Export scoring results as XLSX file.
    """
    session = _get_session(session_id, current_user.tenant.id)
    scoring_result = session.get("scoring_result")
    if scoring_result is None:
        raise HTTPException(
            status_code=400,
            detail="No scoring results. Run /score first.",
        )

    params = session.get("params", ScoringParams())
    text_part_result = session.get("text_part_result")
    campaign_analysis = session.get("campaign_analysis")
    buf = export_results(scoring_result, params, text_part_result=text_part_result, campaign_analysis=campaign_analysis)

    # Audit log
    from database import async_session
    async with async_session() as db:
        await log_audit(
            db, current_user.tenant.id, current_user.user.id, "export", request,
            resource_type="session", resource_id=session_id,
        )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=scoring_results.xlsx"
        },
    )


@app.get("/template")
def download_template():
    """
    Serve the scoring_template.xlsx file.
    """
    template_path = os.path.join(os.path.dirname(__file__), "scoring_template.xlsx")
    if not os.path.exists(template_path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        headers = [
            "text_id", "headline", "campaign", "platform", "device",
            "date_from", "date_to", "impressions", "clicks", "spend",
            "installs", "event_1", "event_2", "event_3", "event_4", "revenue",
            "banner_url",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example rows showing daily granularity format
        example_font = Font(color="808080", italic=True)
        examples = [
            ["ad_001", "Пример заголовка 1", "Кампания A", "google", "mobile",
             "2025-01-01", "2025-01-01", 1500, 45, 2000, 12, 5, 2, 1, 0, 3500,
             "https://cdn.example.com/banner_001.png"],
            ["ad_001", "Пример заголовка 1", "Кампания A", "google", "mobile",
             "2025-01-02", "2025-01-02", 1800, 52, 2200, 15, 7, 3, 1, 0, 4200,
             "https://cdn.example.com/banner_001.png"],
            ["ad_002", "Пример заголовка 2", "Кампания B", "yandex", "desktop",
             "2025-01-01", "2025-01-01", 900, 30, 1500, 8, 3, 1, 0, 0, 2000,
             "https://cdn.example.com/banner_002.jpg"],
        ]
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=scoring_template.xlsx"
            },
        )

    with open(template_path, "rb") as f:
        content = f.read()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=scoring_template.xlsx"
        },
    )


@app.delete("/session/{session_id}")
async def delete_session(session_id: str, current_user: CurrentUser = Depends(get_current_user)):
    """
    Delete session data from memory and soft-delete in DB.
    """
    with _session_lock:
        SESSION_STORE.pop(session_id, None)
        timer = SESSION_TIMERS.pop(session_id, None)
        if timer:
            timer.cancel()

    # Soft-delete in DB
    try:
        from database import async_session
        async with async_session() as db:
            sid = uuid_mod.UUID(session_id)
            await db.execute(
                update(ScoringSession)
                .where(ScoringSession.id == sid)
                .values(status=SessionStatus.deleted)
            )
            await db.commit()
    except Exception as e:
        logger.error("Failed to soft-delete session %s in DB: %s", session_id, e)

    logger.info("Session %s deleted", session_id)
    return {"status": "deleted"}


@app.post("/session/{session_id}/restore")
async def restore_session(
    session_id: str,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Restore a completed session into memory from DB + stored file,
    so it can be re-scored with different parameters.
    """
    tid = current_user.tenant.id
    sid = uuid_mod.UUID(session_id)

    # Check if already in memory
    with _session_lock:
        if session_id in SESSION_STORE:
            s = SESSION_STORE[session_id]
            return {
                "status": "ok",
                "session_id": session_id,
                "mode": s.get("mode"),
                "events": s.get("events", []),
                "columns_detected": s.get("columns_detected", []),
                "auto_mapped": s.get("auto_mapped", {}),
                "n_rows": len(s["df_original"]) if s.get("df_original") is not None else 0,
            }

    # Load session metadata from DB
    result = await db.execute(
        select(ScoringSession)
        .where(ScoringSession.id == sid, ScoringSession.tenant_id == tid)
    )
    session_db = result.scalar_one_or_none()
    if not session_db:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session_db.file_id:
        raise HTTPException(status_code=400, detail="No stored file for this session — re-upload required")

    # Load stored file metadata
    file_result = await db.execute(
        select(StoredFile).where(StoredFile.id == session_db.file_id)
    )
    stored_file = file_result.scalar_one_or_none()
    if not stored_file or stored_file.status != FileStatus.ready:
        raise HTTPException(status_code=400, detail="Stored file not available")

    # Download file content from storage
    try:
        content = await file_storage.download_file(stored_file.storage_key)
    except Exception as e:
        logger.error("Failed to download stored file %s: %s", stored_file.storage_key, e)
        raise HTTPException(status_code=500, detail="Failed to retrieve stored file")

    # Parse file
    try:
        buf = io.BytesIO(content)
        fname = stored_file.original_name.lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf, engine="openpyxl")
    except Exception as e:
        logger.error("Failed to parse restored file: %s", e)
        raise HTTPException(status_code=500, detail="Failed to parse stored file")

    # Restore session metadata
    # Prefer events from params (user may have edited labels on mapping step)
    columns_detected = session_db.columns_detected or list(df.columns.astype(str))
    auto_mapped = session_db.auto_mapped or {}
    mapping = session_db.mapping or auto_mapped
    params_events = (session_db.params or {}).get("events", [])
    events = params_events if params_events else (session_db.events_detected or [])
    mode = session_db.mode

    # Re-apply mapping to get df_mapped
    df_mapped = mapper.apply_mapping(df.copy(), mapping)
    for ev in events:
        col = ev.get("column")
        slot = ev.get("slot")
        if col and slot and col in df_mapped.columns and col != slot:
            df_mapped = df_mapped.rename(columns={col: slot})

    # Store in memory
    with _session_lock:
        SESSION_STORE[session_id] = {
            "tenant_id": tid,
            "df_original": df,
            "df_mapped": df_mapped,
            "columns_detected": columns_detected,
            "auto_mapped": auto_mapped,
            "mapping": mapping,
            "events": events,
            "mode": mode,
            "scoring_result": None,
            "params": session_db.params,
            "text_part_result": None,
        }
        _schedule_ttl_locked(session_id)

    logger.info("Session %s restored from DB (%d rows)", session_id, len(df))
    return {
        "status": "ok",
        "session_id": session_id,
        "mode": mode,
        "events": events,
        "columns_detected": columns_detected,
        "auto_mapped": auto_mapped,
        "n_rows": len(df),
    }
