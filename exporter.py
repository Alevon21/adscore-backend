"""XLSX export with sheets: Results, Statistics, Parameters, Text Parts, and Campaign Analysis."""

import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import ScoringResult, ScoringParams, TextPartAnalysisResult, CampaignAnalysisResult

logger = logging.getLogger(__name__)

# Category colors
CATEGORY_FILLS = {
    "TOP": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "GOOD": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "AVERAGE": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "LOW": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def export_results(
    scoring_result: ScoringResult,
    params: ScoringParams,
    text_part_result: Optional[TextPartAnalysisResult] = None,
    campaign_analysis: Optional[CampaignAnalysisResult] = None,
) -> io.BytesIO:
    """
    Create XLSX with 3+ sheets and return as BytesIO buffer.
    """
    wb = Workbook()

    # ---- Sheet 1: Results ----
    ws_results = wb.active
    ws_results.title = "Результаты"

    # Determine which metrics are available
    available_metrics = set()
    for r in scoring_result.results:
        available_metrics.update(r.metrics.keys())
    ordered_metrics = sorted(available_metrics)

    # Check if any result has a campaign
    has_campaign = any(r.campaign for r in scoring_result.results)

    headers = ["text_id", "headline"]
    if has_campaign:
        headers.append("campaign")
    headers += ["composite_score", "category", "alt_category"]
    headers += ordered_metrics
    headers += [f"z_{m}" for m in ordered_metrics]
    headers += ["n_impressions", "n_clicks", "verdict", "warnings"]

    # Write header
    for col_idx, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Write data
    for row_idx, result in enumerate(scoring_result.results, 2):
        data = [
            result.text_id,
            result.headline[:200] if result.headline else "",
        ]
        if has_campaign:
            data.append(result.campaign or "")
        data += [
            result.composite_score,
            result.category,
            result.alt_category,
        ]
        for m in ordered_metrics:
            val = result.metrics.get(m)
            data.append(round(val, 6) if val is not None else "")
        for m in ordered_metrics:
            val = result.z_scores.get(m)
            data.append(round(val, 4) if val is not None else "")
        data.append(result.n_impressions)
        data.append(result.n_clicks)
        data.append(result.verdict.verdict if result.verdict else "")
        data.append(", ".join(result.warnings) if result.warnings else "")

        cat_col = 4 + (1 if has_campaign else 0)
        for col_idx, value in enumerate(data, 1):
            cell = ws_results.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == cat_col:
                cat_fill = CATEGORY_FILLS.get(str(value))
                if cat_fill:
                    cell.fill = cat_fill

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws_results.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(scoring_result.results) + 2, 102))
        )
        ws_results.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws_results.freeze_panes = "A2"

    # ---- Sheet 2: Statistics ----
    ws_stats = wb.create_sheet("Статистика")
    stats = scoring_result.stats

    stat_rows = [
        ("Параметр", "Значение"),
        ("Всего текстов", stats.get("n_total", 0)),
        ("Оценено", stats.get("n_scored", 0)),
        ("Исключено", stats.get("n_excluded", 0)),
        ("Категория TOP", stats.get("n_top", 0)),
        ("Категория LOW", stats.get("n_low", 0)),
        ("Режим", stats.get("mode", "")),
        ("Средний score", stats.get("score_mean", 0)),
        ("Std score", stats.get("score_std", 0)),
    ]

    for row_idx, (label, value) in enumerate(stat_rows, 1):
        cell_a = ws_stats.cell(row=row_idx, column=1, value=label)
        cell_b = ws_stats.cell(row=row_idx, column=2, value=value)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        if row_idx == 1:
            cell_a.fill = HEADER_FILL
            cell_a.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
            cell_b.font = HEADER_FONT

    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 20

    # ---- Sheet 3: Parameters ----
    ws_params = wb.create_sheet("Параметры")

    param_rows = [
        ("Параметр", "Значение", "Описание"),
        ("alpha", params.alpha, "Уровень значимости"),
        ("fdr_level", params.fdr_level, "Целевой уровень FDR"),
        ("min_impressions", params.min_impressions, "Минимум показов"),
        ("min_clicks", params.min_clicks, "Минимум кликов"),
        ("min_conversions", params.min_conversions, "Минимум конверсий"),
        ("winsorize_lower", params.winsorize_lower, "Нижний перцентиль винсоризации"),
        ("winsorize_upper", params.winsorize_upper, "Верхний перцентиль винсоризации"),
        ("top_threshold", params.top_threshold, "Порог TOP"),
        ("low_threshold", params.low_threshold, "Порог LOW"),
        ("score_quality_floor", params.score_quality_floor, "Мин. балл для TOP"),
        ("good_quality_floor", params.good_quality_floor, "Мин. балл для GOOD"),
    ]

    for metric, weight in params.weights.items():
        param_rows.append((f"weight_{metric}", weight, f"Вес метрики {metric}"))

    for ev in (params.events or []):
        param_rows.append((f"event_{ev.slot}", ev.label, f"Первичное: {'Да' if ev.is_primary else 'Нет'}"))

    for row_idx, (label, value, desc) in enumerate(param_rows, 1):
        cell_a = ws_params.cell(row=row_idx, column=1, value=label)
        cell_b = ws_params.cell(row=row_idx, column=2, value=value)
        cell_c = ws_params.cell(row=row_idx, column=3, value=desc)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        cell_c.border = THIN_BORDER
        if row_idx == 1:
            cell_a.fill = HEADER_FILL
            cell_a.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
            cell_b.font = HEADER_FONT
            cell_c.fill = HEADER_FILL
            cell_c.font = HEADER_FONT

    ws_params.column_dimensions["A"].width = 25
    ws_params.column_dimensions["B"].width = 15
    ws_params.column_dimensions["C"].width = 40

    # ---- Sheet 4: Text Parts Analysis (if available) ----
    if text_part_result and text_part_result.parts_detected:
        ws_parts = wb.create_sheet("Анализ элементов текста")

        part_headers = [
            "Элемент", "Метрика", "N с элементом", "N без", "Среднее с",
            "Среднее без", "Дельта", "Дельта %", "p-value", "Значимо"
        ]
        for col_idx, header in enumerate(part_headers, 1):
            cell = ws_parts.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        row_idx = 2
        for metric_name, impacts in text_part_result.part_impacts.items():
            for impact in impacts:
                data = [
                    impact.part_name,
                    metric_name,
                    impact.n_with,
                    impact.n_without,
                    impact.metric_with,
                    impact.metric_without,
                    impact.delta,
                    impact.delta_pct,
                    impact.p_value,
                    "Да" if impact.significant else "Нет",
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws_parts.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                    if col_idx == 7 and isinstance(value, (int, float)):
                        cell.fill = POSITIVE_FILL if value > 0 else NEGATIVE_FILL
                row_idx += 1

        row_idx += 2
        combo_header_row = row_idx
        combo_headers = ["Ранг", "Элементы", "N текстов", "Средний балл"]
        for col_idx, header in enumerate(combo_headers, 1):
            cell = ws_parts.cell(row=combo_header_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        for combo in text_part_result.best_combinations:
            row_idx += 1
            data = [
                combo.rank,
                " + ".join(combo.parts),
                combo.n_texts,
                combo.avg_metric,
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws_parts.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

        for col_idx in range(1, len(part_headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws_parts.column_dimensions[col_letter].width = 18

        ws_parts.freeze_panes = "A2"

    # ---- Sheet 5: Campaign Analysis (if available) ----
    if campaign_analysis and campaign_analysis.campaigns:
        ws_camp = wb.create_sheet("Анализ кампаний")

        # Determine campaign metrics
        camp_metrics = set()
        for c in campaign_analysis.campaigns:
            camp_metrics.update(c.metrics.keys())
        ordered_camp_metrics = sorted(camp_metrics)

        camp_headers = [
            "Кампания", "Текстов", "Показы", "Клики", "Расход", "Выручка",
            "Балл кампании", "Средний балл текстов", "Категория", "Вердикт",
            "Потери бюджета %", "Разброс качества",
            "Лучший текст", "Балл лучшего", "Худший текст", "Балл худшего",
        ]
        camp_headers += ordered_camp_metrics
        camp_headers += [f"z_{m}" for m in ordered_camp_metrics]

        for col_idx, header in enumerate(camp_headers, 1):
            cell = ws_camp.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        for row_idx, c in enumerate(campaign_analysis.campaigns, 2):
            data = [
                c.campaign,
                c.n_texts_scored or c.n_texts,
                c.total_impressions,
                c.total_clicks,
                round(c.total_spend, 2) if c.total_spend else 0,
                round(c.total_revenue, 2) if c.total_revenue else 0,
                round(c.composite_score, 4),
                round(c.avg_text_score, 4),
                c.category,
                c.verdict.verdict if c.verdict else "",
                round(c.budget_waste_pct, 2),
                round(c.score_spread, 4),
                c.best_text_id,
                round(c.best_text_score, 4),
                c.worst_text_id,
                round(c.worst_text_score, 4),
            ]
            for m in ordered_camp_metrics:
                val = c.metrics.get(m)
                data.append(round(val, 6) if val is not None else "")
            for m in ordered_camp_metrics:
                val = c.z_scores.get(m)
                data.append(round(val, 4) if val is not None else "")

            for col_idx, value in enumerate(data, 1):
                cell = ws_camp.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if col_idx == 9:  # Category column
                    cat_fill = CATEGORY_FILLS.get(str(value))
                    if cat_fill:
                        cell.fill = cat_fill

        for col_idx in range(1, len(camp_headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws_camp.cell(row=r, column=col_idx).value or ""))
                for r in range(1, min(len(campaign_analysis.campaigns) + 2, 102))
            )
            ws_camp.column_dimensions[col_letter].width = min(max_len + 4, 50)

        ws_camp.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
